#!/usr/bin/env python3
"""
experiment_benchmark.py – Three-way comparison: encoding_route vs optimised_encoding
vs encoding_alternative (ClinGcon CSP hybrid), across experimental series and solver
configurations using the Clingo / ClinGcon Python API.

Defaults to the hard series:
  D  Bin scaling          L=6, P=2, TR=2, bins ∈ {1,2,3}
  E  Capacity tightness   L=6, P=2, TR=2, bins=1 (tight/medium/loose)

Each (series, instance, encoding, bins) combination is run under all
solver configurations:
  jumpy, tweety, handy, parallel, parallel_raw

Metrics collected per run
--------------------------
  atoms, rules          grounded program size
  choices, conflicts,   search statistics
  restarts
  ground_time_s         wall-clock time for ctl.ground()
  solve_time_s          wall-clock time for solve (= total − ground)
  total_time_s          total wall-clock time
  result                OPTIMUM | SAT_TIMEOUT | UNSAT | TIMEOUT | ERROR

Note: clingo's Python API (5.8) does not expose a "times" key in
ctl.statistics, so all timing is measured via wall-clock.

Usage
-----
    uv run experiment_benchmark.py                        # series D & E, all configs
    uv run experiment_benchmark.py --series A B C D E    # all original series
    uv run experiment_benchmark.py --time-limit 120      # 2-minute per-run limit
    uv run experiment_benchmark.py --output-dir my_results
    uv run experiment_benchmark.py --threads 8           # parallel config thread count
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import statistics
import sys
import threading
import time
from collections import defaultdict
from dataclasses import asdict, dataclass, fields
from pathlib import Path
from typing import Optional

import json
import subprocess

import clingo
import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── paths ─────────────────────────────────────────────────────────────────────
HERE     = Path(__file__).parent
DATA_DIR = HERE / "data"

ENCODINGS: dict[str, Path] = {
    "encoding_route":                HERE / "encodings" / "encoding_route.lp",
    "optimised_encoding":            HERE / "encodings" / "optimised_encoding.lp",
    "encoding_alternative":          HERE / "encodings" / "encoding_alternative.lp",
    "optimised_encoding_alternative": HERE / "encodings" / "optimised_encoding_alternative.lp",
}

# Solver type per encoding.  "clingcon" encodings are run via an isolated
# subprocess (_clingcon_worker.py); "clingo" encodings run in-process.
ENCODING_TYPE: dict[str, str] = {
    "encoding_route":                "clingo",
    "optimised_encoding":            "clingo",
    "encoding_alternative":          "clingcon",
    "optimised_encoding_alternative": "clingcon",
}

# Adapter injected when solving with encoding_alternative.
# Instances declare transportCapacity/2 but the encoding uses transportResource/1.
_CLINGCON_ADAPTER = "transportResource(TR) :- transportCapacity(TR, _)."

# Series configuration
#   series_dir : subdirectory under DATA_DIR containing the .lp instances
#   bins       : list of bin counts to test (each becomes a separate run)
SERIES_CONFIG: dict[str, dict] = {
    "A": {
        "description": "Location scaling  (P=1, TR=2, bins=1)",
        "series_dir":  "series_A",
        "bins":        [1],
    },
    "B": {
        "description": "Product scaling  (L=6, TR=2, bins=1)",
        "series_dir":  "series_B",
        "bins":        [1],
    },
    "C": {
        "description": "Transport type scaling  (L=6, P=2, bins=1)",
        "series_dir":  "series_C",
        "bins":        [1],
    },
    "D": {
        "description": "Bin scaling  (L=6, P=2, TR=2, bins=1–3)",
        "series_dir":  "series_D",
        "bins":        [1, 2, 3],
    },
    "E": {
        "description": "Capacity tightness  (L=6, P=2, TR=2, bins=1)",
        "series_dir":  "series_E",
        "bins":        [1],
    },
    "F": {
        "description": "Location × bin scaling  (P=1, TR=2, bins=1–3)",
        "series_dir":  "series_A",
        "bins":        [1, 2, 3],
    },
    "G": {
        "description": "Product × bin scaling  (L=6, TR=2, bins=1–3)",
        "series_dir":  "series_B",
        "bins":        [1, 2, 3],
    },
    "H": {
        "description": "Transport × bin scaling  (L=6, P=2, bins=1–3)",
        "series_dir":  "series_C",
        "bins":        [1, 2, 3],
    },
    "I": {
        "description": "Capacity tightness × bin scaling  (L=6, P=2, TR=2, bins=1–3)",
        "series_dir":  "series_E",
        "bins":        [1, 2, 3],
    },
}

# Solver configurations to sweep.
# "parallel" is built at runtime to include the user-specified thread count.
_BASE_CONFIGURATIONS: dict[str, list[str]] = {
    "jumpy":  ["--configuration=jumpy"],
    "tweety": ["--configuration=tweety"],
    "handy":  ["--configuration=handy"],
}


def build_configurations(num_threads: int) -> dict[str, list[str]]:
    """Return all configurations, including two parallel variants:
      parallel     -- --configuration=many (portfolio preset) + threads
      parallel_raw -- threads only, clingo's default heuristic
    """
    configs = dict(_BASE_CONFIGURATIONS)
    configs["parallel"]     = ["--configuration=many", f"--parallel-mode={num_threads}"]
    configs["parallel_raw"] = [f"--parallel-mode={num_threads}"]
    return configs


# ── result record ─────────────────────────────────────────────────────────────

@dataclass
class RunRecord:
    timestamp:     str
    series:        str
    instance:      str
    encoding:      str
    bins:          int
    configuration: str
    result:        str          # OPTIMUM | SAT_TIMEOUT | UNSAT | TIMEOUT | ERROR
    # Grounded program
    atoms:         Optional[int]
    rules:         Optional[int]
    # Search statistics
    choices:       Optional[int]
    conflicts:     Optional[int]
    restarts:      Optional[int]
    # Timing (seconds, all wall-clock)
    ground_time_s: Optional[float]
    solve_time_s:  Optional[float]
    total_time_s:  Optional[float]


# ── statistics helpers ────────────────────────────────────────────────────────

def _get(d, *keys, default=None):
    """Safely traverse a nested clingo stats object."""
    v = d
    for k in keys:
        try:
            v = v[k]
        except (KeyError, TypeError, IndexError):
            return default
    return v


def _extract_stats(ctl: clingo.Control, wall_ground: float, wall_total: float) -> dict:
    """
    Pull grounding/search stats from ctl.statistics and timing from wall-clock.

    clingo's Python API (5.8) does not expose a "times" key in ctl.statistics,
    so all timing is wall-clock only:
      ground_time_s = time spent in ctl.ground()
      total_time_s  = ground + solve (wall-clock)
      solve_time_s  = total − ground
    """
    s = ctl.statistics

    atoms = _get(s, "problem", "lp", "atoms")
    rules = _get(s, "problem", "lp", "rules")

    # clingo reports search stats per solver thread; sum across threads
    solvers_raw = _get(s, "solving", "solvers", default={})
    if isinstance(solvers_raw, list):
        choices   = sum(int(_get(sv, "choices",   default=0)) for sv in solvers_raw)
        conflicts = sum(int(_get(sv, "conflicts",  default=0)) for sv in solvers_raw)
        restarts  = sum(int(_get(sv, "restarts",   default=0)) for sv in solvers_raw)
    else:
        choices   = _get(solvers_raw, "choices",   default=None)
        conflicts = _get(solvers_raw, "conflicts",  default=None)
        restarts  = _get(solvers_raw, "restarts",   default=None)

    solve_t = max(wall_total - wall_ground, 0.0)

    return dict(
        atoms         = int(atoms)         if atoms    is not None else None,
        rules         = int(rules)         if rules    is not None else None,
        choices       = int(choices)       if choices  is not None else None,
        conflicts     = int(conflicts)     if conflicts is not None else None,
        restarts      = int(restarts)      if restarts is not None else None,
        ground_time_s = round(wall_ground, 4),
        solve_time_s  = round(solve_t,     4),
        total_time_s  = round(wall_total,  4),
    )


# ── solver ────────────────────────────────────────────────────────────────────

def run_instance(
    encoding_path: Path,
    instance_path: Path,
    num_bins:      int,
    time_limit:    int,
    config_args:   list[str],
) -> dict:
    """
    Solve one (encoding, instance, bins, configuration) combination.

    bins(1..num_bins) is injected so that both encodings see the same bin set.
    A threading.Timer enforces the wall-clock time limit by calling handle.cancel().
    """
    ctl = clingo.Control([
        "--stats=2",
        f"-c", f"numBins={num_bins}",
        *config_args,
    ])

    ctl.load(str(encoding_path))
    ctl.load(str(instance_path))

    # Provide bins for encoding_route (which has no bins rule);
    # optimised_encoding's bins(1..numBins) gives the same set — union is identical.
    ctl.add("base", [], f"bins(1..{num_bins}).")

    t0 = time.perf_counter()
    ctl.ground([("base", [])])
    wall_ground = time.perf_counter() - t0

    has_model      = False
    optimum_proven = False

    with ctl.solve(yield_=True) as handle:
        timer = threading.Timer(time_limit, handle.cancel)
        timer.start()
        try:
            for model in handle:
                has_model = True
                if model.optimality_proven:
                    optimum_proven = True
        finally:
            timer.cancel()
        solve_result = handle.get()

    wall_total = time.perf_counter() - t0

    interrupted = getattr(solve_result, "interrupted", not solve_result.exhausted)
    if solve_result.satisfiable is False:
        result = "UNSAT"
    elif optimum_proven:
        result = "OPTIMUM"
    elif has_model and interrupted:
        result = "SAT_TIMEOUT"
    elif interrupted:
        result = "TIMEOUT"
    else:
        result = "SATISFIABLE"   # model found, search exhausted, no optimality flag

    stats = _extract_stats(ctl, wall_ground, wall_total)
    stats["result"] = result
    return stats


# ── ClinGcon helpers ──────────────────────────────────────────────────────────

def _get_instance_bounds(instance_path: Path) -> tuple[int, int]:
    """Return (max_freq, max_items) by parsing the numFreq/numParts atoms in
    the instance file.  These become the maxFreq and maxItems constants required
    by encoding_alternative.lp's theory domains."""
    ctl = clingo.Control()
    ctl.load(str(instance_path))
    ctl.ground([("base", [])])
    max_freq = max_items = 0
    for atom in ctl.symbolic_atoms:
        sym = atom.symbol
        if sym.name == "numFreq" and len(sym.arguments) == 1:
            max_freq  = max(max_freq,  sym.arguments[0].number)
        elif sym.name == "numParts" and len(sym.arguments) == 1:
            max_items = max(max_items, sym.arguments[0].number)
    return max_freq, max_items


_WORKER = HERE / "_clingcon_worker.py"


def run_instance_clingcon(
    encoding_path: Path,
    instance_path: Path,
    num_bins:      int,
    time_limit:    int,
    config_args:   list[str],
    max_freq:      int,
    max_items:     int,
) -> dict:
    """Solve one combination using the ClinGcon theory API in an isolated subprocess.

    ClinGcon can segfault on certain (bins, configuration) combinations.
    Running it as a subprocess means a crash kills only the child process;
    the benchmark runner catches the non-zero exit code and records ERROR.

    The child (_clingcon_worker.py) prints a single JSON object to stdout.
    A subprocess timeout of time_limit + 30 s gives the child enough slack to
    report SAT_TIMEOUT / TIMEOUT itself before we force-kill it.
    """
    _ERROR_STATS = dict(
        result="ERROR", atoms=None, rules=None,
        choices=None, conflicts=None, restarts=None,
        ground_time_s=None, solve_time_s=None, total_time_s=None,
    )

    cmd = [
        sys.executable, str(_WORKER),
        str(encoding_path), str(instance_path),
        str(num_bins), str(time_limit), str(max_freq), str(max_items),
        *config_args,
    ]

    try:
        proc = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=time_limit + 30,   # hard kill after solver's own limit + grace
        )
    except subprocess.TimeoutExpired:
        return {**_ERROR_STATS, "result": "TIMEOUT"}

    if proc.returncode != 0:
        # Segfault (139), OOM, or other fatal error
        raise RuntimeError(
            f"clingcon worker exited with code {proc.returncode}"
            + (f": {proc.stderr.strip()[:200]}" if proc.stderr.strip() else "")
        )

    return json.loads(proc.stdout.strip())


# ── benchmark runner ──────────────────────────────────────────────────────────

def run_benchmarks(
    series_names: list[str],
    time_limit:   int,
    output_dir:   Path,
    num_threads:  int,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    configurations = build_configurations(num_threads)

    all_records: list[RunRecord] = []
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")

    # Count total runs upfront for progress display
    total_runs = 0
    for s in series_names:
        cfg   = SERIES_CONFIG[s]
        insts = list((DATA_DIR / cfg["series_dir"]).glob("*.lp"))
        total_runs += len(insts) * len(ENCODINGS) * len(cfg["bins"]) * len(configurations)

    run_n = 0

    for series in series_names:
        cfg        = SERIES_CONFIG[series]
        series_dir = DATA_DIR / cfg["series_dir"]
        instances  = sorted(series_dir.glob("*.lp"))

        if not instances:
            print(f"WARNING: no instances in {series_dir} — did you run generate_instances.py?",
                  file=sys.stderr)
            continue

        print(f"\n{'─'*70}")
        print(f"Series {series}: {cfg['description']}")
        print(f"{'─'*70}")

        for inst_path in instances:
            # Pre-compute clingcon bounds once per instance (cheap; separate ctl)
            inst_max_freq, inst_max_items = _get_instance_bounds(inst_path)

            for enc_name, enc_path in ENCODINGS.items():
                for bins in cfg["bins"]:
                    for cfg_name, cfg_args in configurations.items():
                        run_n += 1
                        tag = (f"[{run_n:>3}/{total_runs}] "
                               f"enc={enc_name:<22} bins={bins} "
                               f"cfg={cfg_name:<10} inst={inst_path.name}")
                        print(f"  {tag}", end=" … ", flush=True)

                        try:
                            if ENCODING_TYPE[enc_name] == "clingcon":
                                stats = run_instance_clingcon(
                                    enc_path, inst_path, bins, time_limit, cfg_args,
                                    inst_max_freq, inst_max_items,
                                )
                            else:
                                stats = run_instance(
                                    enc_path, inst_path, bins, time_limit, cfg_args,
                                )
                        except Exception as exc:
                            print(f"ERROR: {exc}")
                            stats = dict(
                                result        = "ERROR",
                                atoms         = None, rules      = None,
                                choices       = None, conflicts  = None,
                                restarts      = None,
                                ground_time_s = None,
                                solve_time_s  = None,
                                total_time_s  = None,
                            )

                        res = stats["result"]
                        t   = stats["total_time_s"]
                        ch  = stats["choices"]
                        print(f"{res:<12}  {f'{t:.3f}s' if t is not None else '-':>8}"
                              f"  choices={ch if ch is not None else '-'}")

                        all_records.append(RunRecord(
                            timestamp     = dt.datetime.now().isoformat(),
                            series        = series,
                            instance      = inst_path.name,
                            encoding      = enc_name,
                            bins          = bins,
                            configuration = cfg_name,
                            **stats,
                        ))

    # ── write CSV ─────────────────────────────────────────────────────────────
    csv_path = output_dir / f"results_{timestamp}.csv"
    fieldnames = [f.name for f in fields(RunRecord)]
    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for rec in all_records:
            writer.writerow(asdict(rec))

    print(f"\nResults saved → {csv_path}")

    # ── Excel export ───────────────────────────────────────────────────────────
    xlsx_path = csv_path.with_suffix(".xlsx")
    export_excel(all_records, xlsx_path)

    # ── print summary ─────────────────────────────────────────────────────────
    _print_summary(all_records)

    return csv_path


# ── summary helpers ────────────────────────────────────────────────────────────

def _avg(vals: list) -> Optional[float]:
    clean = [v for v in vals if v is not None]
    return statistics.mean(clean) if clean else None


def _fmt(v, fmt=".3f") -> str:
    return f"{v:{fmt}}" if v is not None else "-"


def _print_summary(records: list[RunRecord]) -> None:
    print(f"\n{'═'*115}")
    print("SUMMARY")
    print(f"{'═'*115}")

    groups: dict[tuple, list[RunRecord]] = defaultdict(list)
    for r in records:
        groups[(r.series, r.encoding, r.bins, r.configuration)].append(r)

    col = "{:<6} {:<22} {:>4} {:<12} {:>4} {:>7} {:>9} {:>9} {:>11} {:>10}"
    header = col.format(
        "Series", "Encoding", "bins", "Config",
        "N", "OPT", "avg_t(s)", "avg_s(s)", "avg_choices", "avg_atoms",
    )
    print(header)
    print("─" * 115)

    for (series, enc, bins, cfg), grp in sorted(groups.items()):
        n_opt  = sum(1 for r in grp if r.result == "OPTIMUM")
        avg_t  = _avg([r.total_time_s  for r in grp])
        avg_s  = _avg([r.solve_time_s  for r in grp])
        avg_ch = _avg([r.choices       for r in grp])
        avg_at = _avg([r.atoms         for r in grp])

        print(col.format(
            series, enc, bins, cfg, len(grp), n_opt,
            _fmt(avg_t,  ".3f"),
            _fmt(avg_s,  ".3f"),
            _fmt(avg_ch, ".0f"),
            _fmt(avg_at, ".0f"),
        ))

    print(f"{'═'*115}")

    # Per (series, bins, config): pairwise solve-time ratios
    print("\nKey ratios (solve_time: lower ratio = numerator faster)")
    print("─" * 90)
    by_key: dict[tuple, dict[str, list[RunRecord]]] = defaultdict(lambda: defaultdict(list))
    for r in records:
        by_key[(r.series, r.bins, r.configuration)][r.encoding].append(r)

    for (series, bins, cfg), enc_map in sorted(by_key.items()):
        base = enc_map.get("encoding_route", [])
        opt  = enc_map.get("optimised_encoding", [])
        alt  = enc_map.get("encoding_alternative", [])

        def _rat(num_recs, den_recs, attr):
            d = _avg([getattr(r, attr) for r in den_recs])
            n = _avg([getattr(r, attr) for r in num_recs])
            return f"{n/d:.2f}x" if d and n and d > 0 else "-"

        print(f"  S{series} bins={bins} {cfg:<12}:  "
              f"opt/base={_rat(opt, base, 'solve_time_s')}  "
              f"alt/base={_rat(alt, base, 'solve_time_s')}  "
              f"alt/opt={_rat(alt, opt, 'solve_time_s')}")


# ── Excel styling helpers ─────────────────────────────────────────────────────

_RAW_COLS     = [f.name for f in fields(RunRecord)]
_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_SUBHEAD_FILL = PatternFill("solid", fgColor="2E75B6")
_OPT_FILL     = PatternFill("solid", fgColor="C6EFCE")
_BASE_FILL    = PatternFill("solid", fgColor="FCE4D6")
_ALT_FILL     = PatternFill("solid", fgColor="DDEBF7")
_OPT_ALT_FILL = PatternFill("solid", fgColor="9DC3E6")
_WHITE_FONT   = Font(color="FFFFFF", bold=True)
_BOLD_FONT    = Font(bold=True)
_HEAT_COLS    = {"choices", "conflicts", "restarts",
                 "total_time_s", "solve_time_s", "ground_time_s",
                 "atoms", "rules"}


def _set_col_widths(ws, widths: dict[int, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _apply_heatmap(ws, headers: list[str], data_start_row: int, max_row: int) -> None:
    if max_row < data_start_row:
        return
    for ci, name in enumerate(headers, 1):
        if name in _HEAT_COLS:
            col_letter = get_column_letter(ci)
            ws.conditional_formatting.add(
                f"{col_letter}{data_start_row}:{col_letter}{max_row}",
                ColorScaleRule(
                    start_type="min",      start_color="63BE7B",
                    mid_type="percentile", mid_value=50, mid_color="FFEB84",
                    end_type="max",        end_color="F8696B",
                ),
            )


# ── Excel sheets ──────────────────────────────────────────────────────────────

def _write_raw_sheet(wb: openpyxl.Workbook, records: list[RunRecord]) -> None:
    ws = wb.active
    ws.title = "Raw Data"
    for ci, name in enumerate(_RAW_COLS, 1):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _WHITE_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    for ri, rec in enumerate(records, 2):
        row = asdict(rec)
        for ci, name in enumerate(_RAW_COLS, 1):
            ws.cell(row=ri, column=ci, value=row[name])
    _apply_heatmap(ws, _RAW_COLS, 2, len(records) + 1)
    # timestamp(1) series(2) instance(3) encoding(4) bins(5) configuration(6)
    # result(7) atoms(8) rules(9) choices(10) conflicts(11) restarts(12)
    # ground(13) solve(14) total(15)
    _set_col_widths(ws, {1: 26, 2: 8, 3: 30, 4: 22, 5: 5, 6: 12,
                         7: 14, 8: 9, 9: 9, 10: 10, 11: 10, 12: 9,
                         13: 13, 14: 13, 15: 13})


def _write_summary_sheet(wb: openpyxl.Workbook, records: list[RunRecord]) -> None:
    """One row per (series, encoding, bins, configuration) with averaged metrics."""
    ws = wb.create_sheet("Summary")
    headers = [
        "series", "description", "encoding", "bins", "configuration",
        "N", "OPTIMUM", "TIMEOUT",
        "avg_total_s", "avg_ground_s", "avg_solve_s",
        "avg_choices", "avg_conflicts", "avg_restarts",
        "avg_atoms", "avg_rules",
    ]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _WHITE_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    groups: dict[tuple, list[RunRecord]] = defaultdict(list)
    for r in records:
        groups[(r.series, r.encoding, r.bins, r.configuration)].append(r)

    ri = 2
    for (series, enc, bins, cfg), grp in sorted(groups.items()):
        n_timeout = sum(1 for r in grp if r.result in ("TIMEOUT", "SAT_TIMEOUT"))
        n_opt     = sum(1 for r in grp if r.result == "OPTIMUM")
        desc      = SERIES_CONFIG.get(series, {}).get("description", "")
        if enc == "optimised_encoding":
            fill = _OPT_FILL
        elif enc == "optimised_encoding_alternative":
            fill = _OPT_ALT_FILL
        elif enc == "encoding_alternative":
            fill = _ALT_FILL
        else:
            fill = _BASE_FILL

        row_vals = [
            series, desc, enc, bins, cfg, len(grp), n_opt, n_timeout,
            _avg([r.total_time_s  for r in grp]),
            _avg([r.ground_time_s for r in grp]),
            _avg([r.solve_time_s  for r in grp]),
            _avg([r.choices       for r in grp]),
            _avg([r.conflicts     for r in grp]),
            _avg([r.restarts      for r in grp]),
            _avg([r.atoms         for r in grp]),
            _avg([r.rules         for r in grp]),
        ]
        for ci, v in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci,
                           value=round(v, 4) if isinstance(v, float) else v)
            cell.fill = fill
        ri += 1

    _apply_heatmap(ws, headers, 2, ri - 1)
    _set_col_widths(ws, {1: 8, 2: 45, 3: 22, 4: 5, 5: 12,
                         6: 4, 7: 8, 8: 8,
                         9: 12, 10: 12, 11: 12,
                         12: 12, 13: 12, 14: 10, 15: 10, 16: 10})


def _write_ratios_sheet(wb: openpyxl.Workbook, records: list[RunRecord]) -> None:
    """Three-way speedup ratios per (series, bins, configuration).

    Columns show opt/base, alt/base, and alt/opt ratios for solve time,
    choices, and total time, plus raw averages for all three encodings.
    Green cell = numerator faster (ratio < 1); red = numerator slower.
    """
    ws = wb.create_sheet("Ratios (3-way)")
    headers = [
        "series", "bins", "configuration", "description",
        # opt ÷ base
        "opt/base_solve_t", "opt/base_total_t", "opt/base_choices",
        # alt ÷ base
        "alt/base_solve_t", "alt/base_total_t", "alt/base_choices",
        # alt ÷ opt
        "alt/opt_solve_t",  "alt/opt_total_t",  "alt/opt_choices",
        # raw averages
        "base_avg_total_s", "opt_avg_total_s", "alt_avg_total_s",
        "base_avg_choices", "opt_avg_choices", "alt_avg_choices",
    ]
    RATIO_COLS = set(range(5, 14))   # 1-based columns containing ratio values

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _WHITE_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    by_key: dict[tuple, dict[str, list[RunRecord]]] = defaultdict(lambda: defaultdict(list))
    for r in records:
        by_key[(r.series, r.bins, r.configuration)][r.encoding].append(r)

    ri = 2
    for (series, bins, cfg), enc_map in sorted(by_key.items()):
        base = enc_map.get("encoding_route", [])
        opt  = enc_map.get("optimised_encoding", [])
        alt  = enc_map.get("encoding_alternative", [])

        def _ratio(numerator_recs, denominator_recs, attr):
            d = _avg([getattr(r, attr) for r in denominator_recs])
            n = _avg([getattr(r, attr) for r in numerator_recs])
            return round(n / d, 4) if d and n and d > 0 else None

        desc = SERIES_CONFIG.get(series, {}).get("description", "")
        row_vals = [
            series, bins, cfg, desc,
            _ratio(opt,  base, "solve_time_s"),
            _ratio(opt,  base, "total_time_s"),
            _ratio(opt,  base, "choices"),
            _ratio(alt,  base, "solve_time_s"),
            _ratio(alt,  base, "total_time_s"),
            _ratio(alt,  base, "choices"),
            _ratio(alt,  opt,  "solve_time_s"),
            _ratio(alt,  opt,  "total_time_s"),
            _ratio(alt,  opt,  "choices"),
            _avg([r.total_time_s for r in base]),
            _avg([r.total_time_s for r in opt]),
            _avg([r.total_time_s for r in alt]),
            _avg([r.choices      for r in base]),
            _avg([r.choices      for r in opt]),
            _avg([r.choices      for r in alt]),
        ]
        for ci, v in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci,
                           value=round(v, 4) if isinstance(v, float) else v)
            if ci in RATIO_COLS and isinstance(v, float):
                if v < 1.0:
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                elif v > 1.0:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")
        ri += 1

    _set_col_widths(ws, {1: 8, 2: 5, 3: 12, 4: 45,
                         5: 16, 6: 16, 7: 16,
                         8: 16, 9: 16, 10: 16,
                         11: 15, 12: 15, 13: 15,
                         14: 17, 15: 16, 16: 16,
                         17: 17, 18: 16, 19: 16})


def _write_config_comparison_sheet(wb: openpyxl.Workbook, records: list[RunRecord]) -> None:
    """
    One sheet per (series, bins): configurations as column groups, instances as rows.

    Layout:
      instance | encoding | default_total | default_solve | default_choices | default_result
                           | frumpy_total  | frumpy_solve  | frumpy_choices  | frumpy_result
                           | ...
    This makes it easy to see which configuration performs best for each instance.
    """
    by_series_bins: dict[tuple, list[RunRecord]] = defaultdict(list)
    for r in records:
        by_series_bins[(r.series, r.bins)].append(r)

    all_configs = sorted({r.configuration for r in records})

    for (series, bins), recs in sorted(by_series_bins.items()):
        ws = wb.create_sheet(f"Config S{series} bins={bins}")
        desc = SERIES_CONFIG.get(series, {}).get("description", "")
        ws.cell(row=1, column=1,
                value=f"Series {series} bins={bins}: {desc} — configuration comparison").font = _BOLD_FONT

        base_headers = ["instance", "encoding"]
        config_headers: list[str] = []
        for cfg in all_configs:
            config_headers += [
                f"{cfg}_total_s",
                f"{cfg}_solve_s",
                f"{cfg}_choices",
                f"{cfg}_result",
            ]
        all_headers = base_headers + config_headers

        for ci, h in enumerate(all_headers, 1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.fill = _SUBHEAD_FILL
            cell.font = _WHITE_FONT
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "C3"

        # Index records by (instance, encoding, configuration)
        idx: dict[tuple, RunRecord] = {}
        for r in recs:
            idx[(r.instance, r.encoding, r.configuration)] = r

        instances = sorted({r.instance for r in recs})
        encodings = sorted(ENCODINGS.keys())

        ri = 3
        for inst in instances:
            for enc in encodings:
                if enc == "optimised_encoding":
                    fill = _OPT_FILL
                elif enc == "encoding_alternative":
                    fill = _ALT_FILL
                else:
                    fill = _BASE_FILL
                row_vals: list = [inst, enc]
                for cfg in all_configs:
                    rec = idx.get((inst, enc, cfg))
                    if rec:
                        row_vals += [rec.total_time_s, rec.solve_time_s,
                                     rec.choices, rec.result]
                    else:
                        row_vals += [None, None, None, None]

                for ci, v in enumerate(row_vals, 1):
                    ws.cell(row=ri, column=ci, value=v).fill = fill
                ri += 1

        _set_col_widths(ws, {1: 30, 2: 22})
        for ci in range(3, len(all_headers) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 13


def _write_per_series_sheet(wb: openpyxl.Workbook, series: str,
                             records: list[RunRecord]) -> None:
    """Detailed sheet for one series: all records sorted by instance/encoding/config."""
    ws = wb.create_sheet(f"Series {series}")
    desc = SERIES_CONFIG.get(series, {}).get("description", "")
    ws.cell(row=1, column=1, value=f"Series {series}: {desc}").font = _BOLD_FONT

    headers = ["instance", "bins", "encoding", "configuration", "result",
               "total_time_s", "solve_time_s", "ground_time_s",
               "choices", "conflicts", "restarts", "atoms", "rules"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.fill = _SUBHEAD_FILL
        cell.font = _WHITE_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A3"

    series_recs = [r for r in records if r.series == series]
    for ri, rec in enumerate(series_recs, 3):
        row_vals = [rec.instance, rec.bins, rec.encoding, rec.configuration,
                    rec.result, rec.total_time_s, rec.solve_time_s,
                    rec.ground_time_s, rec.choices, rec.conflicts,
                    rec.restarts, rec.atoms, rec.rules]
        if rec.encoding == "optimised_encoding":
            fill = _OPT_FILL
        elif rec.encoding == "optimised_encoding_alternative":
            fill = _OPT_ALT_FILL
        elif rec.encoding == "encoding_alternative":
            fill = _ALT_FILL
        else:
            fill = _BASE_FILL
        for ci, v in enumerate(row_vals, 1):
            ws.cell(row=ri, column=ci, value=v).fill = fill

    _apply_heatmap(ws, headers, 3, len(series_recs) + 2)
    _set_col_widths(ws, {1: 30, 2: 5, 3: 22, 4: 12, 5: 14,
                         6: 13, 7: 13, 8: 13,
                         9: 10, 10: 10, 11: 9, 12: 10, 13: 9})


def export_excel(records: list[RunRecord], path: Path) -> None:
    """Write a multi-sheet Excel workbook analysing all benchmark records."""
    wb = openpyxl.Workbook()
    _write_raw_sheet(wb, records)
    _write_summary_sheet(wb, records)
    _write_ratios_sheet(wb, records)
    _write_config_comparison_sheet(wb, records)
    for s in sorted({r.series for r in records}):
        _write_per_series_sheet(wb, s, records)
    wb.save(path)
    print(f"Excel workbook saved → {path}")


# ── CLI ───────────────────────────────────────────────────────────────────────

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Benchmark encoding_route vs optimised_encoding across experiment "
            "series and solver configurations using the Clingo Python API."
        )
    )
    parser.add_argument(
        "--series", nargs="+",
        choices=sorted(SERIES_CONFIG),
        default=["D", "E"],
        help="Series to run (default: D E — the hard series).",
    )
    parser.add_argument(
        "--time-limit", type=int, default=300,
        help="Per-run wall-clock time limit in seconds (default: 300).",
    )
    parser.add_argument(
        "--output-dir", default="benchmark_results",
        help="Directory for CSV output (default: benchmark_results/).",
    )
    parser.add_argument(
        "--threads", type=int, default=4,
        help="Thread count for the 'parallel' configuration (default: 4).",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()

    for enc_name, path in ENCODINGS.items():
        if not path.exists():
            print(f"Error: encoding not found ({enc_name}): {path}", file=sys.stderr)
            return 1

    run_benchmarks(
        series_names = args.series,
        time_limit   = args.time_limit,
        output_dir   = HERE / args.output_dir,
        num_threads  = args.threads,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
