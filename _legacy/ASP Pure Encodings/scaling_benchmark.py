#!/usr/bin/env python3
"""
scaling_benchmark.py — Three-encoding scaling comparison.

Generates (or reuses) Series F and G instances, then benchmarks three encodings
at bins=2 with a single solver configuration:

  encoding_route               — baseline pure ASP
  optimised_encoding           — optimised pure ASP
  optimised_encoding_alternative — ClinGcon CSP hybrid

Series
------
  F  Location scaling   L ∈ {4,6,8,10,12,15}, P=2, TR=2, bins=2
  G  Product scaling    L=8, P ∈ {1,2,3,4,5}, TR=2, bins=2

Usage
-----
    uv run scaling_benchmark.py                         # generate + run both series
    uv run scaling_benchmark.py --no-generate           # skip instance generation
    uv run scaling_benchmark.py --series F              # only location scaling
    uv run scaling_benchmark.py --time-limit 120
    uv run scaling_benchmark.py --config jumpy          # jumpy | tweety | handy
    uv run scaling_benchmark.py --output-dir my_results
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import subprocess
import sys
from collections import defaultdict
from dataclasses import dataclass, fields, asdict
from pathlib import Path
from typing import Optional

import time

import clingo
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

# ── paths ─────────────────────────────────────────────────────────────────────
HERE     = Path(__file__).parent
DATA_DIR = HERE / "data"

ENCODINGS: dict[str, Path] = {
    "encoding_route":                 HERE / "encodings" / "encoding_route.lp",
    "optimised_encoding":             HERE / "encodings" / "optimised_encoding.lp",
    "optimised_encoding_alternative": HERE / "encodings" / "optimised_encoding_alternative.lp",
}

ENCODING_TYPE: dict[str, str] = {
    "encoding_route":                 "clingo",
    "optimised_encoding":             "clingo",
    "optimised_encoding_alternative": "clingcon",
}

_CLINGO_WORKER   = HERE / "_clingo_worker.py"
_CLINGCON_WORKER = HERE / "_clingcon_worker.py"

# Excel fill colours per encoding
_BASE_FILL = PatternFill("solid", fgColor="FCE4D6")   # light orange
_OPT_FILL  = PatternFill("solid", fgColor="E2EFDA")   # light green
_ALT_FILL  = PatternFill("solid", fgColor="9DC3E6")   # blue

ENCODING_FILL = {
    "encoding_route":                 _BASE_FILL,
    "optimised_encoding":             _OPT_FILL,
    "optimised_encoding_alternative": _ALT_FILL,
}

# Series definitions: {series_name: (dir_name, instance_glob, description)}
SERIES_DEF: dict[str, tuple[str, str, str]] = {
    "F": ("series_F", "L*_P02_TR2_s*.lp", "Location scaling  (P=2, TR=2, bins=2)"),
    "G": ("series_G", "L08_P*_TR2_s*.lp", "Product scaling  (L=8, TR=2, bins=2)"),
}

NUM_BINS    = 2
DEFAULT_TL  = 300   # seconds per run


# ── result record ─────────────────────────────────────────────────────────────

@dataclass
class RunRecord:
    timestamp:     str
    series:        str
    instance:      str
    encoding:      str
    bins:          int
    configuration: str
    result:        str
    atoms:         Optional[int]
    rules:         Optional[int]
    choices:       Optional[int]
    conflicts:     Optional[int]
    restarts:      Optional[int]
    ground_time_s: Optional[float]
    solve_time_s:  Optional[float]
    total_time_s:  Optional[float]


# ── statistics helpers ─────────────────────────────────────────────────────────

def _get(d, *keys, default=None):
    v = d
    for k in keys:
        try:
            v = v[k]
        except (KeyError, TypeError, IndexError):
            return default
    return v


# ── pure-ASP runner (subprocess-isolated to handle SIGBUS/segfault) ───────────

def run_clingo(
    encoding_path: Path,
    instance_path: Path,
    num_bins:      int,
    time_limit:    int,
    config_args:   list[str],
) -> dict:
    cmd = [
        sys.executable, str(_CLINGO_WORKER),
        str(encoding_path), str(instance_path),
        str(num_bins), str(time_limit),
        *config_args,
    ]
    try:
        proc = subprocess.run(
            cmd, capture_output=True, text=True, timeout=time_limit + 30
        )
    except subprocess.TimeoutExpired:
        return dict(_ERROR_STATS, result="TIMEOUT")

    if proc.returncode != 0:
        print(f"      [clingo worker error rc={proc.returncode}]", flush=True)
        if proc.stderr:
            print(f"      stderr: {proc.stderr[:200]}", flush=True)
        return dict(_ERROR_STATS)

    try:
        return json.loads(proc.stdout.strip())
    except json.JSONDecodeError:
        return dict(_ERROR_STATS)


# ── ClinGcon runner (subprocess-isolated) ─────────────────────────────────────

def _get_instance_bounds(instance_path: Path) -> tuple[int, int]:
    ctl = clingo.Control()
    ctl.load(str(instance_path))
    ctl.ground([("base", [])])
    max_freq = max_items = 0
    for atom in ctl.symbolic_atoms:
        sym = atom.symbol
        if sym.name == "numFreq" and len(sym.arguments) == 1:
            max_freq  = max(max_freq,  sym.arguments[0].number)
        elif sym.name == "numParts" and len(sym.arguments) == 1:
            max_items = max(max_items, sym.arguments[0].number)
    return max_freq, max_items


_ERROR_STATS = dict(
    result="ERROR", atoms=None, rules=None,
    choices=None, conflicts=None, restarts=None,
    ground_time_s=None, solve_time_s=None, total_time_s=None,
)


def run_clingcon(
    encoding_path: Path,
    instance_path: Path,
    num_bins:      int,
    time_limit:    int,
    config_args:   list[str],
    max_freq:      int,
    max_items:     int,
) -> dict:
    cmd = [
        sys.executable, str(_CLINGCON_WORKER),
        str(encoding_path), str(instance_path),
        str(num_bins), str(time_limit), str(max_freq), str(max_items),
        *config_args,
    ]
    try:
        proc = subprocess.run(
            cmd, capture_output=True, text=True, timeout=time_limit + 30
        )
    except subprocess.TimeoutExpired:
        return dict(_ERROR_STATS, result="TIMEOUT")

    if proc.returncode != 0:
        print(f"      [worker error rc={proc.returncode}]", flush=True)
        if proc.stderr:
            print(f"      stderr: {proc.stderr[:200]}", flush=True)
        return dict(_ERROR_STATS)

    try:
        return json.loads(proc.stdout.strip())
    except json.JSONDecodeError:
        return dict(_ERROR_STATS)


# ── main runner ───────────────────────────────────────────────────────────────

def run_all(
    series_list:  list[str],
    time_limit:   int,
    config_name:  str,
    config_args:  list[str],
    output_dir:   Path,
) -> list[RunRecord]:
    records: list[RunRecord] = []
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")

    for series_name in series_list:
        dir_name, glob_pat, desc = SERIES_DEF[series_name]
        series_dir = DATA_DIR / dir_name

        if not series_dir.exists():
            print(f"  [SKIP] series {series_name}: directory not found ({series_dir})")
            continue

        instances = sorted(series_dir.glob(glob_pat))
        if not instances:
            print(f"  [SKIP] series {series_name}: no instances matching '{glob_pat}'")
            continue

        print(f"\nSeries {series_name} — {desc}")
        print(f"  {len(instances)} instances × {len(ENCODINGS)} encodings × bins={NUM_BINS}")

        for inst_path in instances:
            max_freq, max_items = _get_instance_bounds(inst_path)

            for enc_name, enc_path in ENCODINGS.items():
                enc_type = ENCODING_TYPE[enc_name]
                print(f"    {inst_path.name}  enc={enc_name}  bins={NUM_BINS} ...",
                      end=" ", flush=True)
                t_start = time.perf_counter()

                try:
                    if enc_type == "clingo":
                        stats = run_clingo(
                            enc_path, inst_path, NUM_BINS, time_limit, config_args
                        )
                    else:
                        stats = run_clingcon(
                            enc_path, inst_path, NUM_BINS, time_limit, config_args,
                            max_freq, max_items
                        )
                except Exception as exc:
                    print(f"ERROR: {exc}", flush=True)
                    stats = dict(_ERROR_STATS)

                elapsed = time.perf_counter() - t_start
                res = stats.get("result", "ERROR")
                gt  = stats.get("ground_time_s")
                st  = stats.get("solve_time_s")
                print(
                    f"{res}  ground={gt:.3f}s  solve={st:.3f}s  wall={elapsed:.1f}s"
                    if gt is not None and st is not None
                    else f"{res}  wall={elapsed:.1f}s",
                    flush=True,
                )

                records.append(RunRecord(
                    timestamp     = ts,
                    series        = series_name,
                    instance      = inst_path.name,
                    encoding      = enc_name,
                    bins          = NUM_BINS,
                    configuration = config_name,
                    result        = res,
                    atoms         = stats.get("atoms"),
                    rules         = stats.get("rules"),
                    choices       = stats.get("choices"),
                    conflicts     = stats.get("conflicts"),
                    restarts      = stats.get("restarts"),
                    ground_time_s = stats.get("ground_time_s"),
                    solve_time_s  = stats.get("solve_time_s"),
                    total_time_s  = stats.get("total_time_s"),
                ))

    return records


# ── CSV output ────────────────────────────────────────────────────────────────

def write_csv(records: list[RunRecord], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fnames = [f.name for f in fields(RunRecord)]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fnames)
        w.writeheader()
        w.writerows(asdict(r) for r in records)
    print(f"\nCSV written: {path}")


# ── Excel output ──────────────────────────────────────────────────────────────

_HEADER_FONT  = Font(bold=True)
_CENTER       = Alignment(horizontal="center")
_TIMEOUT_FILL = PatternFill("solid", fgColor="FFD966")   # yellow
_ERROR_FILL   = PatternFill("solid", fgColor="FF0000")   # red


def _cell(ws, row, col, value, *, bold=False, fill=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    if bold:
        c.font = Font(bold=True)
    if fill:
        c.fill = fill
    if align:
        c.alignment = align
    return c


def write_excel(records: list[RunRecord], path: Path) -> None:
    wb = openpyxl.Workbook()

    enc_names = list(ENCODINGS.keys())

    # ── per-series sheets ──────────────────────────────────────────────────────
    for series_name in sorted({r.series for r in records}):
        sr = [r for r in records if r.series == series_name]
        ws = wb.create_sheet(title=f"Series {series_name}")

        # Column layout: instance | (ground, solve, total, result) × encoding
        cols_per_enc = ["ground_time_s", "solve_time_s", "total_time_s", "result"]
        header_row1  = ["Instance"]
        header_row2  = [""]
        for enc in enc_names:
            header_row1 += [enc] + [""] * (len(cols_per_enc) - 1)
            header_row2 += cols_per_enc

        # Write two-level header
        for ci, h in enumerate(header_row1, 1):
            _cell(ws, 1, ci, h, bold=True, fill=ENCODING_FILL.get(
                enc_names[(ci - 2) // len(cols_per_enc)]
                if ci > 1 else None, None
            ), align=_CENTER)
        for ci, h in enumerate(header_row2, 1):
            _cell(ws, 2, ci, h, bold=True, align=_CENTER)

        # Merge header cells per encoding
        for ei, enc in enumerate(enc_names):
            start = 2 + ei * len(cols_per_enc)
            end   = start + len(cols_per_enc) - 1
            ws.merge_cells(
                start_row=1, start_column=start,
                end_row=1,   end_column=end
            )
            ws.cell(1, start).fill  = ENCODING_FILL[enc]
            ws.cell(1, start).font  = Font(bold=True)
            ws.cell(1, start).alignment = _CENTER

        # Group by instance
        instances = sorted({r.instance for r in sr})
        row = 3
        for inst in instances:
            _cell(ws, row, 1, inst)
            for ei, enc in enumerate(enc_names):
                recs = [r for r in sr if r.instance == inst and r.encoding == enc]
                if not recs:
                    col_start = 2 + ei * len(cols_per_enc)
                    for ci in range(len(cols_per_enc)):
                        _cell(ws, row, col_start + ci, "—")
                    continue
                rec = recs[0]   # bins=2 fixed, one config
                col_start = 2 + ei * len(cols_per_enc)
                for ci, field in enumerate(cols_per_enc):
                    val  = getattr(rec, field)
                    cell = _cell(ws, row, col_start + ci, val, fill=ENCODING_FILL[enc])
                    if field == "result":
                        if val in ("TIMEOUT", "SAT_TIMEOUT"):
                            cell.fill = _TIMEOUT_FILL
                        elif val == "ERROR":
                            cell.fill = _ERROR_FILL
                    if isinstance(val, float):
                        cell.number_format = "0.000"
            row += 1

        # Auto-width
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value is not None),
                default=8,
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    # ── summary sheet: median total_time per (series, instance_group, encoding) ─
    ws_sum = wb.create_sheet(title="Summary", index=0)
    _cell(ws_sum, 1, 1, "Scaling Benchmark Summary — bins=2", bold=True)

    # Build summary: for each series, group instances by their "scale parameter"
    # (L or P, extracted from filename) and show median total_time
    row = 3
    for series_name in sorted({r.series for r in records}):
        _cell(ws_sum, row, 1, f"Series {series_name}", bold=True)
        row += 1

        # Header
        _cell(ws_sum, row, 1, "Instance", bold=True)
        for ei, enc in enumerate(enc_names):
            _cell(ws_sum, row, 2 + ei * 3,     "ground_s", bold=True, fill=ENCODING_FILL[enc])
            _cell(ws_sum, row, 2 + ei * 3 + 1, "solve_s",  bold=True, fill=ENCODING_FILL[enc])
            _cell(ws_sum, row, 2 + ei * 3 + 2, "result",   bold=True, fill=ENCODING_FILL[enc])
        row += 1

        sr = [r for r in records if r.series == series_name]
        for inst in sorted({r.instance for r in sr}):
            _cell(ws_sum, row, 1, inst)
            for ei, enc in enumerate(enc_names):
                recs = [r for r in sr if r.instance == inst and r.encoding == enc]
                if not recs:
                    for ci in range(3):
                        _cell(ws_sum, row, 2 + ei * 3 + ci, "—")
                    continue
                rec = recs[0]
                gt_cell = _cell(ws_sum, row, 2 + ei * 3,
                                rec.ground_time_s, fill=ENCODING_FILL[enc])
                st_cell = _cell(ws_sum, row, 2 + ei * 3 + 1,
                                rec.solve_time_s, fill=ENCODING_FILL[enc])
                res_cell = _cell(ws_sum, row, 2 + ei * 3 + 2,
                                 rec.result, fill=ENCODING_FILL[enc])
                if isinstance(rec.ground_time_s, float):
                    gt_cell.number_format = "0.000"
                if isinstance(rec.solve_time_s, float):
                    st_cell.number_format = "0.000"
                if rec.result in ("TIMEOUT", "SAT_TIMEOUT"):
                    res_cell.fill = _TIMEOUT_FILL
                elif rec.result == "ERROR":
                    res_cell.fill = _ERROR_FILL
            row += 1
        row += 1   # blank separator

    # Auto-width summary sheet
    for col in ws_sum.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=8,
        )
        ws_sum.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    # Remove default blank sheet if present
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Excel written: {path}")


# ── terminal summary ──────────────────────────────────────────────────────────

def print_summary(records: list[RunRecord]) -> None:
    print("\n" + "=" * 70)
    print("SCALING BENCHMARK SUMMARY  (bins=2)")
    print("=" * 70)

    enc_names = list(ENCODINGS.keys())

    for series_name in sorted({r.series for r in records}):
        _, _, desc = SERIES_DEF[series_name]
        print(f"\nSeries {series_name} — {desc}")

        # Column widths
        instances = sorted({r.instance for r in records if r.series == series_name})
        hdr = f"  {'Instance':<28}" + "".join(f"  {e[-20:]:<22}" for e in enc_names)
        print(hdr)
        print("  " + "-" * (len(hdr) - 2))

        for inst in instances:
            row = f"  {inst:<28}"
            for enc in enc_names:
                recs = [r for r in records
                        if r.series == series_name and r.instance == inst
                        and r.encoding == enc]
                if not recs:
                    row += f"  {'—':<22}"
                    continue
                rec = recs[0]
                if rec.total_time_s is not None:
                    cell = f"{rec.result:<10} {rec.total_time_s:>6.2f}s"
                else:
                    cell = f"{rec.result:<10} {'N/A':>7}"
                row += f"  {cell:<22}"
            print(row)

    # Speedup ratios: opt/base and clingcon/base
    print("\n  Speedup ratios  (total_time: base → optimised, base → clingcon)")
    print("  " + "-" * 60)
    for series_name in sorted({r.series for r in records}):
        instances = sorted({r.instance for r in records if r.series == series_name})
        for inst in instances:
            def get_t(enc):
                recs = [r for r in records
                        if r.series == series_name and r.instance == inst
                        and r.encoding == enc and r.total_time_s is not None]
                return recs[0].total_time_s if recs else None

            t_base = get_t("encoding_route")
            t_opt  = get_t("optimised_encoding")
            t_alt  = get_t("optimised_encoding_alternative")

            if t_base and t_opt and t_alt:
                print(f"  {inst:<28}  opt/base={t_opt/t_base:.2f}x  "
                      f"clingcon/base={t_alt/t_base:.2f}x")


# ── instance generation ───────────────────────────────────────────────────────

def generate_instances() -> None:
    """Run generate_instances.py to produce series F and G."""
    gen_script = HERE / "generate_instances.py"
    print("Generating series F and G instances …")
    result = subprocess.run(
        [sys.executable, str(gen_script)],
        capture_output=False,
    )
    if result.returncode != 0:
        print("WARNING: instance generation returned non-zero exit code", file=sys.stderr)


# ── CLI ────────────────────────────────────────────────────────────────────────

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Three-encoding scaling benchmark at bins=2"
    )
    p.add_argument(
        "--series", nargs="+", choices=["F", "G"], default=["F", "G"],
        help="Which series to run (default: F G)"
    )
    p.add_argument(
        "--no-generate", action="store_true",
        help="Skip instance generation (use existing data/series_F and data/series_G)"
    )
    p.add_argument(
        "--time-limit", type=int, default=DEFAULT_TL,
        help=f"Per-run time limit in seconds (default: {DEFAULT_TL})"
    )
    p.add_argument(
        "--config", default="jumpy",
        choices=["jumpy", "tweety", "handy", "default"],
        help="Clingo solver configuration (default: jumpy)"
    )
    p.add_argument(
        "--output-dir", type=Path, default=HERE / "benchmark_results",
        help="Directory for CSV and Excel output"
    )
    return p


def main() -> None:
    args = build_parser().parse_args()

    # Instance generation
    if not args.no_generate:
        generate_instances()

    # Solver configuration
    config_name = args.config
    config_args = [f"--configuration={config_name}"] if config_name != "default" else []

    print(f"\nBenchmark settings:")
    print(f"  Series     : {args.series}")
    print(f"  Bins       : {NUM_BINS}")
    print(f"  Config     : {config_name}")
    print(f"  Time limit : {args.time_limit}s per run")
    print(f"  Encodings  : {list(ENCODINGS.keys())}")

    # Run
    records = run_all(
        series_list  = args.series,
        time_limit   = args.time_limit,
        config_name  = config_name,
        config_args  = config_args,
        output_dir   = args.output_dir,
    )

    if not records:
        print("\nNo records collected — check that instances exist.")
        return

    # Output
    ts = records[0].timestamp
    csv_path   = args.output_dir / f"scaling_{ts}.csv"
    excel_path = args.output_dir / f"scaling_{ts}.xlsx"

    write_csv(records, csv_path)
    write_excel(records, excel_path)
    print_summary(records)


if __name__ == "__main__":
    main()
