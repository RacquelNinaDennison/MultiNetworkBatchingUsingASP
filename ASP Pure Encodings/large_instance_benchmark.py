#!/usr/bin/env python3
"""
large_instance_benchmark.py — benchmark alternative_clingcon.lp on the 20
large instances from nonmonotonic-generator/generated_instances_large/.

These instances have 51 locations (17 harbors, 34 prodlocs), 35 parts, and
~5096 routes.  They use industry-scale predicates (harbor, transportCO2, etc.)
which match the structure assumed by alternative_clingcon.lp.

Adapters injected automatically:
  - partTR(P,TR) :- part(P), transportCapacity(TR,_).
  - transportResource(TR) :- transportCapacity(TR,_).
  - cap_size_divide overridden to 1 (instances have raw capacities, not millions)

Usage
-----
    uv run large_instance_benchmark.py
    uv run large_instance_benchmark.py --time-limit 300
    uv run large_instance_benchmark.py --config jumpy
    uv run large_instance_benchmark.py --instances 1 2 3   # specific instance numbers
    uv run large_instance_benchmark.py --output-dir my_results
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import subprocess
import sys
import time
from dataclasses import asdict, dataclass, fields
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── paths ─────────────────────────────────────────────────────────────────────
HERE         = Path(__file__).parent
REPO_ROOT    = HERE.parent
INSTANCES_DIR = REPO_ROOT / "nonmonotonic-generator" / "generated_instances_large"
ENCODING_PATH = HERE / "encodings" / "alternative_clingcon.lp"
WORKER        = HERE / "_large_clingcon_worker.py"

DEFAULT_TL   = 300   # seconds per run
ALL_INSTANCES = list(range(1, 21))


# ── result record ─────────────────────────────────────────────────────────────

@dataclass
class RunRecord:
    timestamp:     str
    instance:      str
    configuration: str
    result:        str
    atoms:         Optional[int]
    rules:         Optional[int]
    choices:       Optional[int]
    conflicts:     Optional[int]
    restarts:      Optional[int]
    ground_time_s: Optional[float]
    solve_time_s:  Optional[float]
    total_time_s:  Optional[float]


# ── runner ────────────────────────────────────────────────────────────────────

_ERROR_STATS = dict(
    result="ERROR", atoms=None, rules=None,
    choices=None, conflicts=None, restarts=None,
    ground_time_s=None, solve_time_s=None, total_time_s=None,
)


def run_instance(
    instance_path: Path,
    time_limit:    int,
    config_name:   str,
    config_args:   list[str],
) -> dict:
    """Run one instance via the isolated subprocess worker."""
    cmd = [
        sys.executable, str(WORKER),
        str(ENCODING_PATH),
        str(instance_path),
        str(time_limit),
        "-c", "cap_size_divide=1",   # override: instances use raw capacity values
        *config_args,
    ]
    try:
        proc = subprocess.run(
            cmd, capture_output=True, text=True, timeout=time_limit + 30
        )
    except subprocess.TimeoutExpired:
        return dict(_ERROR_STATS, result="TIMEOUT")

    if proc.returncode != 0:
        print(f"      [worker error rc={proc.returncode}]", flush=True)
        if proc.stderr:
            # Print first 300 chars of stderr for diagnosis
            print(f"      stderr: {proc.stderr[:300]}", flush=True)
        return dict(_ERROR_STATS)

    try:
        return json.loads(proc.stdout.strip())
    except json.JSONDecodeError:
        return dict(_ERROR_STATS)


def run_all(
    instance_nums: list[int],
    time_limit:    int,
    config_name:   str,
    config_args:   list[str],
) -> list[RunRecord]:
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    records: list[RunRecord] = []

    print(f"\nEncoding : {ENCODING_PATH.name}")
    print(f"Config   : {config_name}")
    print(f"Limit    : {time_limit}s per instance")
    print(f"Instance : {INSTANCES_DIR}")
    print(f"Running  : {len(instance_nums)} instances\n")

    for n in instance_nums:
        inst_path = INSTANCES_DIR / f"instance_large_{n}.lp"
        if not inst_path.exists():
            print(f"  SKIP  {inst_path.name}  (not found)")
            continue

        print(f"  {inst_path.name} ...", end=" ", flush=True)
        t_start = time.perf_counter()

        stats = run_instance(inst_path, time_limit, config_name, config_args)

        wall = time.perf_counter() - t_start
        res  = stats.get("result", "ERROR")
        gt   = stats.get("ground_time_s")
        st   = stats.get("solve_time_s")

        if gt is not None and st is not None:
            print(f"{res:<12}  ground={gt:.3f}s  solve={st:.3f}s  wall={wall:.1f}s",
                  flush=True)
        else:
            print(f"{res:<12}  wall={wall:.1f}s", flush=True)

        records.append(RunRecord(
            timestamp     = ts,
            instance      = inst_path.name,
            configuration = config_name,
            result        = res,
            atoms         = stats.get("atoms"),
            rules         = stats.get("rules"),
            choices       = stats.get("choices"),
            conflicts     = stats.get("conflicts"),
            restarts      = stats.get("restarts"),
            ground_time_s = stats.get("ground_time_s"),
            solve_time_s  = stats.get("solve_time_s"),
            total_time_s  = stats.get("total_time_s"),
        ))

    return records


# ── CSV output ────────────────────────────────────────────────────────────────

def write_csv(records: list[RunRecord], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fnames = [f.name for f in fields(RunRecord)]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fnames)
        w.writeheader()
        w.writerows(asdict(r) for r in records)
    print(f"\nCSV written : {path}")


# ── Excel output ──────────────────────────────────────────────────────────────

_HDR_FONT    = Font(bold=True)
_HDR_FILL    = PatternFill("solid", fgColor="9DC3E6")   # blue (alt encoding colour)
_OPTIMUM_FILL   = PatternFill("solid", fgColor="E2EFDA")   # green
_TIMEOUT_FILL   = PatternFill("solid", fgColor="FFD966")   # yellow
_SAT_FILL       = PatternFill("solid", fgColor="DDEBF7")   # light blue
_ERROR_FILL     = PatternFill("solid", fgColor="FF7575")   # red
_CENTER         = Alignment(horizontal="center")


def _result_fill(result: str) -> Optional[PatternFill]:
    if result == "OPTIMUM":    return _OPTIMUM_FILL
    if result in ("TIMEOUT", "SAT_TIMEOUT"): return _TIMEOUT_FILL
    if result == "SATISFIABLE": return _SAT_FILL
    if result == "ERROR":       return _ERROR_FILL
    return None


def write_excel(records: list[RunRecord], path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Large Instance Results"

    headers = [
        "Instance", "Result",
        "Ground (s)", "Solve (s)", "Total (s)",
        "Atoms", "Rules", "Choices", "Conflicts", "Restarts",
        "Config",
    ]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font  = _HDR_FONT
        cell.fill  = _HDR_FILL
        cell.alignment = _CENTER

    for ri, rec in enumerate(records, 2):
        values = [
            rec.instance, rec.result,
            rec.ground_time_s, rec.solve_time_s, rec.total_time_s,
            rec.atoms, rec.rules, rec.choices, rec.conflicts, rec.restarts,
            rec.configuration,
        ]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if isinstance(val, float):
                cell.number_format = "0.000"
        # Colour-code the result cell
        fill = _result_fill(rec.result)
        if fill:
            ws.cell(row=ri, column=2).fill = fill

    # Auto-width
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Excel written: {path}")


# ── terminal summary ──────────────────────────────────────────────────────────

def print_summary(records: list[RunRecord]) -> None:
    solved   = [r for r in records if r.result == "OPTIMUM"]
    sat_to   = [r for r in records if r.result == "SAT_TIMEOUT"]
    timeouts = [r for r in records if r.result == "TIMEOUT"]
    errors   = [r for r in records if r.result == "ERROR"]

    print("\n" + "=" * 65)
    print(f"LARGE INSTANCE BENCHMARK — alternative_clingcon.lp")
    print("=" * 65)
    print(f"  Total instances : {len(records)}")
    print(f"  OPTIMUM         : {len(solved)}")
    print(f"  SAT_TIMEOUT     : {len(sat_to)}   (solution found, optimality unproven)")
    print(f"  TIMEOUT         : {len(timeouts)}")
    print(f"  ERROR           : {len(errors)}")

    if solved:
        g_times = [r.ground_time_s for r in solved if r.ground_time_s is not None]
        s_times = [r.solve_time_s  for r in solved if r.solve_time_s  is not None]
        t_times = [r.total_time_s  for r in solved if r.total_time_s  is not None]
        print(f"\n  Solved instances — timing:")
        print(f"    ground  min={min(g_times):.2f}s  max={max(g_times):.2f}s  "
              f"avg={sum(g_times)/len(g_times):.2f}s")
        print(f"    solve   min={min(s_times):.2f}s  max={max(s_times):.2f}s  "
              f"avg={sum(s_times)/len(s_times):.2f}s")
        print(f"    total   min={min(t_times):.2f}s  max={max(t_times):.2f}s  "
              f"avg={sum(t_times)/len(t_times):.2f}s")

    print("\n  Per-instance results:")
    print(f"  {'Instance':<28}  {'Result':<12}  {'Ground':>8}  {'Solve':>8}  {'Total':>8}")
    print("  " + "-" * 70)
    for r in records:
        gt = f"{r.ground_time_s:.3f}s" if r.ground_time_s is not None else "N/A"
        st = f"{r.solve_time_s:.3f}s"  if r.solve_time_s  is not None else "N/A"
        tt = f"{r.total_time_s:.3f}s"  if r.total_time_s  is not None else "N/A"
        print(f"  {r.instance:<28}  {r.result:<12}  {gt:>8}  {st:>8}  {tt:>8}")


# ── CLI ────────────────────────────────────────────────────────────────────────

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Benchmark alternative_clingcon.lp on large instances"
    )
    p.add_argument(
        "--instances", nargs="+", type=int, default=ALL_INSTANCES,
        metavar="N",
        help="Instance numbers to run (default: 1–20)"
    )
    p.add_argument(
        "--time-limit", type=int, default=DEFAULT_TL,
        help=f"Per-instance time limit in seconds (default: {DEFAULT_TL})"
    )
    p.add_argument(
        "--config", default="jumpy",
        choices=["jumpy", "tweety", "handy", "default"],
        help="Clingo solver configuration (default: jumpy)"
    )
    p.add_argument(
        "--output-dir", type=Path, default=HERE / "benchmark_results",
        help="Directory for CSV and Excel output"
    )
    return p


def main() -> None:
    args = build_parser().parse_args()

    if not ENCODING_PATH.exists():
        print(f"ERROR: encoding not found: {ENCODING_PATH}", file=sys.stderr)
        sys.exit(1)
    if not INSTANCES_DIR.exists():
        print(f"ERROR: instances directory not found: {INSTANCES_DIR}", file=sys.stderr)
        sys.exit(1)

    config_name = args.config
    config_args = [f"--configuration={config_name}"] if config_name != "default" else []

    records = run_all(
        instance_nums = sorted(set(args.instances)),
        time_limit    = args.time_limit,
        config_name   = config_name,
        config_args   = config_args,
    )

    if not records:
        print("No records — check that instance files exist.")
        return

    ts         = records[0].timestamp
    csv_path   = args.output_dir / f"large_instances_{ts}.csv"
    excel_path = args.output_dir / f"large_instances_{ts}.xlsx"

    write_csv(records, csv_path)
    write_excel(records, excel_path)
    print_summary(records)


if __name__ == "__main__":
    main()
